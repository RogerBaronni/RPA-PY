from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium.common.exceptions import NoSuchElementException

servico = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=servico)

fii = ['KNCR11', 'BCFF11', 'XPLG11', 'HGLG11', 'KNRI11', 'HGRU11', 'VISC11', 'XPML11', 'TRXF11', 'HFOF11', 'BTLG11', 'HGRE11', 'JSRE11', 'HTMX11']
fii.sort()
fiis = list()
açao = ['VALE3', 'ITUB4', 'BBDC4', 'ABEV3', 'PETR4', 'PETR3', 'BBAS3', 'MGLU3', 'BBAS3']
açao.sort()
açoes = list()

list_valor_atual = list()
list_min_52sem = list()
list_min_mes = list()
list_max_52sem = list()
list_max_mes = list()
list_dividend = list()
list_valorizaçao = list()
list_seta_valor_atual = list()
list_seta_valorizaçao = list()
list_itens = list()
num = 0

list_valor_atual1 = list()
list_min_52sem1 = list()
list_min_mes1 = list()
list_max_52sem1 = list()
list_max_mes1 = list()
list_dividend1 = list()
list_valorizaçao1 = list()
list_seta_valor_atual1 = list()
list_seta_valorizaçao1 = list()
list_itens1 = list()
num1 = 0

for i in fii:
    try:
        driver.get(f'https://statusinvest.com.br/fundos-imobiliarios/{i}')
        valor_atual = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[1]/div[1]/div/div[1]/strong')
        list_valor_atual.append(valor_atual.text)

        min_52sem = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[1]/div[2]/div/div[1]/strong')
        list_min_52sem.append(min_52sem.text)

        min_mes = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[1]/div[2]/div/div[2]/div/span[2]')
        list_min_mes.append(min_mes.text)

        max_52sem = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[1]/div[3]/div/div[1]/strong')
        list_max_52sem.append(max_52sem.text)

        max_mes = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[1]/div[3]/div/div[2]/div/span[2]')
        list_max_mes.append(max_mes.text)

        dividend = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[1]/div[4]/div/div[1]/strong')
        list_dividend.append(dividend.text)

        valorizaçao = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[1]/div[5]/div/div[1]/strong')
        list_valorizaçao.append(valorizaçao.text)

        seta_valor_atual = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[1]/div[1]/div/div[2]/span/span/i')
        list_seta_valor_atual.append(seta_valor_atual.text)

        seta_valorizaçao = driver.find_element(By.XPATH, '/html/body/main/div[2]/div[1]/div[5]/div/div[1]/span/i')
        list_seta_valorizaçao.append(seta_valorizaçao.text)

        fiis.append(i)
        num+=1
        list_itens.append(num)

    except NoSuchElementException:
        print(f'fii não encontrado: {i}')

    except Exception as e:
        print(f'erro: {e}')

for i in açao:
    try:
        driver.get(f'https://statusinvest.com.br/acoes/{i}')
        valor_atual1 = driver.find_element(By.XPATH, '/html/body/main/div[2]/div/div[1]/div/div[1]/div/div[1]/strong')
        list_valor_atual1.append(valor_atual1.text)

        min_52sem1 = driver.find_element(By.XPATH, '/html/body/main/div[2]/div/div[1]/div/div[2]/div/div[1]/strong')
        list_min_52sem1.append(min_52sem1.text)

        min_mes1 = driver.find_element(By.XPATH, '/html/body/main/div[2]/div/div[1]/div/div[2]/div/div[2]/div/span[2]')
        list_min_mes1.append(min_mes1.text)

        max_52sem1 = driver.find_element(By.XPATH, '/html/body/main/div[2]/div/div[1]/div/div[3]/div/div[1]/strong')
        list_max_52sem1.append(max_52sem1.text)

        max_mes1 = driver.find_element(By.XPATH, '/html/body/main/div[2]/div/div[1]/div/div[3]/div/div[2]/div/span[2]')
        list_max_mes1.append(max_mes1.text)

        dividend1 = driver.find_element(By.XPATH, '/html/body/main/div[2]/div/div[1]/div/div[4]/div/div[1]/strong')
        list_dividend1.append(dividend1.text)

        valorizaçao1 = driver.find_element(By.XPATH, '/html/body/main/div[2]/div/div[1]/div/div[5]/div/div[1]/strong')
        list_valorizaçao1.append(valorizaçao1.text)

        seta_valor_atual1 = driver.find_element(By.XPATH, '/html/body/main/div[2]/div/div[1]/div/div[1]/div/div[2]/span/span/i')
        list_seta_valor_atual1.append(seta_valor_atual1.text)

        seta_valorizaçao1 = driver.find_element(By.XPATH, '/html/body/main/div[2]/div/div[1]/div/div[5]/div/div[1]/span/i')
        list_seta_valorizaçao1.append(seta_valorizaçao1.text)

        açoes.append(i)
        num1+=1
        list_itens1.append(num1)

    except NoSuchElementException:
        print(f'ação não encontrada: {i}')

    except Exception as e:
        print(f'erro: {e}')

df_fii = pd.DataFrame({'Fiis': fiis} )
df_valor_atual = pd.DataFrame({'Valor atual':list_valor_atual})
df_min_52sem = pd.DataFrame({'Min. 52 semanas': list_min_52sem})
df_min_mes = pd.DataFrame({'Min. mês': list_min_mes})
df_max_52sem = pd.DataFrame({'Max. 52 semanas': list_max_52sem})
df_max_mes = pd.DataFrame({'Max. mês': list_max_mes})
df_dividend = pd.DataFrame({'Dividend Yield': list_dividend})
df_valorizaçao = pd.DataFrame({'Valorização 12M': list_valorizaçao})
df_itens = pd.DataFrame({'Numeração': list_itens})

df_concatenado = pd.concat([df_itens, df_fii, df_valor_atual, df_min_52sem, df_min_mes, df_max_52sem, df_max_mes, df_dividend, df_valorizaçao], axis=1)

caminho_arquivo = 'planilha.xlsx'

df_açao = pd.DataFrame({'Ações': açoes} )
df_valor_atual1 = pd.DataFrame({'Valor atual':list_valor_atual1})
df_min_52sem1 = pd.DataFrame({'Min. 52 semanas': list_min_52sem1})
df_min_mes1 = pd.DataFrame({'Min. mês': list_min_mes1})
df_max_52sem1 = pd.DataFrame({'Max. 52 semanas': list_max_52sem1})
df_max_mes1 = pd.DataFrame({'Max. mês': list_max_mes1})
df_dividend1 = pd.DataFrame({'Dividend Yield': list_dividend1})
df_valorizaçao1 = pd.DataFrame({'Valorização 12M': list_valorizaçao1})
df_itens1 = pd.DataFrame({'Numeração': list_itens1})

df_concatenado1 = pd.concat([df_itens1, df_açao, df_valor_atual1, df_min_52sem1, df_min_mes1, df_max_52sem1, df_max_mes1, df_dividend1, df_valorizaçao1], axis=1)

with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
    df_concatenado.to_excel(writer, sheet_name='FIIS', index=False)
    df_concatenado1.to_excel(writer, sheet_name='Ações', index=False)

wb = load_workbook(caminho_arquivo)
wf = wb['FIIS']
wa = wb['Ações']

col_valor_atual = wf['C']
cor_vermelha = PatternFill(start_color='FF3333', end_color='FF3333', fill_type='solid')
cor_verde = PatternFill(start_color='33FF33', end_color='33FF33', fill_type='solid')
for idx in range(1, len(col_valor_atual), + 1):
    if idx - 1 < len(list_seta_valor_atual):
        cell = wf.cell(row=idx + 1, column=3)
        if list_seta_valor_atual[idx - 1] == 'arrow_upward':
            cell.fill = cor_verde
        elif list_seta_valor_atual[idx - 1] == 'arrow_downward':
            cell.fill = cor_vermelha

col_valorizaçao = wf['I']
for idx in range(1, len(col_valorizaçao), + 1):
    if idx - 1 < len(list_seta_valorizaçao):
        cell1 = wf.cell(row=idx + 1, column=9)
        if list_seta_valorizaçao[idx - 1] == 'arrow_upward':
            cell1.fill = cor_verde
        elif list_seta_valorizaçao[idx - 1] == 'arrow_downward':
            cell1.fill = cor_vermelha

col_valor_atual1 = wa['C']
for idx in range(1, len(col_valor_atual1), + 1):
    if idx - 1 < len(list_seta_valor_atual1):
        cell2 = wa.cell(row=idx + 1, column=3)
        if list_seta_valor_atual1[idx - 1] == 'arrow_upward':
            cell2.fill = cor_verde
        elif list_seta_valor_atual1[idx - 1] == 'arrow_downward':
            cell2.fill = cor_vermelha

col_valorizaçao1 = wa['I']
for idx in range(1, len(col_valorizaçao1), + 1):
    if idx - 1 < len(list_seta_valorizaçao1):
        cell3 = wa.cell(row=idx + 1, column=9)
        if list_seta_valorizaçao1[idx - 1] == 'arrow_upward':
            cell3.fill = cor_verde
        elif list_seta_valorizaçao1[idx - 1] == 'arrow_downward':
            cell3.fill = cor_vermelha

for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    wf.column_dimensions[i].width = 16
    wa.column_dimensions[i].width = 16

wb.save(caminho_arquivo)
